"""从模型 Excel 生成 RTE 读取声明。

脚本会读取 model/SF020_SysFaultMngt.xlsx 中 Rport 工作表的 B 列，
并更新 src/SF020_SysFaultMngt.c 里的 SF020_SysFaultMngt_Cycle_5ms：
1) 为 B 列每个条目生成一条 UINT8 <name>; 局部变量声明。
2) 为每个条目生成一条 (void)Rte_Read_<name>_gdu8(&<name>); 调用。
"""
from pathlib import Path
import re
from openpyxl import load_workbook

# 根据脚本所在目录定位到 [SF020_SysFaultMngt.xlsx](http://_vscodecontentref_/0)
script_dir = Path(__file__).resolve().parent
xlsx_path = script_dir.parent / "model" / "SF020_SysFaultMngt.xlsx"

wb = load_workbook(xlsx_path, data_only=True, read_only=True)
ws = wb["Rport"]

b_column_list = []
for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
    value = row[0]
    if value is not None and str(value).strip():
        b_column_list.append(str(value).strip())

invalid_names = [name for name in b_column_list if not re.match(r"^[A-Za-z_][A-Za-z0-9_]*$", name)]
if invalid_names:
    raise ValueError(f"发现不合法的 C 标识符: {invalid_names}")

wb.close()

print("B列数据：", b_column_list)

# 将 B 列数据按顺序转换为 C 变量声明语句。
ResetIndex_Code = ["  Monr_SF020_u8_SysFaultIndex_UL = 0U;"]
ReturnValue_Code = [f"  boolean value;\n"
  f"  boolean* ret = &value;"]
VariableDeclare_Code = [f"UINT8 {name};" for name in b_column_list]
ReadInput_Code = [f"(void)Rte_Read_{name}_gdu8(&{name});" for name in b_column_list]
VariableDeclare_Code = [f"  UINT8 {name};" for name in b_column_list]
ReadInput_Code = [f"  (void)Rte_Read_{name}_gdu8(&{name});" for name in b_column_list]
FaultTriggerCond_Code = [
    (
        f"  if (({name} >= 1U) && (Monr_SF020_b_SysFaultStatus_UL[Monr_SF020_u8_SysFaultIndex_UL] == 0U))\n"
        "  {\n"
        f"    TriggerFault(Monr_SF020_u8_SysFaultIndex_UL, {name}, ret);\n"
        "  }"
    )
    for name in b_column_list
]
FaultRecoverCond_Code = [
    (
        f"  else if (({name} == 0U) && (Monr_SF020_b_SysFaultStatus_UL[Monr_SF020_u8_SysFaultIndex_UL] == 1U))\n"
        "  {\n"
        "    RecoverFault(Monr_SF020_u8_SysFaultIndex_UL, ret);\n"
        "  }"
    )
    for name in b_column_list
]
ElseCond_Code = [
    "  else\n"
    "  {\n"
    "    /* no action */\n"
    "  }"
    for _ in b_column_list
]

VariableUpdate_Code = [f"  Monr_SF020_u8_SysFaultIndex_UL++;" for _ in b_column_list]
BlockStart = "  /* CODEGEN_MERGED_CODE_START */"
BlockEnd = "  /* CODEGEN_MERGED_CODE_END */"


# 按要求顺序拼接代码：
# ResetIndex_Code -> VariableDeclare_Code -> ReadInput_Code ->
# (FaultTriggerCond_Code, FaultRecoverCond_Code, ElseCond_Code， VariableUpdate_Code) 按元素依次拼接
Merged_Code = []
Merged_Code.extend(ResetIndex_Code)
Merged_Code.extend(ReturnValue_Code)
Merged_Code.append("")
Merged_Code.extend(VariableDeclare_Code)
Merged_Code.append("")
Merged_Code.extend(ReadInput_Code)
Merged_Code.append("")

# 遍历所有故障信号，为每一个生成完整的触发/恢复逻辑块
for idx in range(len(b_column_list)):
    Merged_Code.append(FaultTriggerCond_Code[idx])
    Merged_Code.append(FaultRecoverCond_Code[idx])
    Merged_Code.append(ElseCond_Code[idx])
    Merged_Code.append(VariableUpdate_Code[idx])
    Merged_Code.append("")


def insert_code_at_function_start(c_text: str, func_signature: str, new_body_lines: list[str]) -> str:
    signature_index = c_text.find(func_signature)
    if signature_index == -1:
        raise ValueError(f"未找到函数签名: {func_signature}")

    brace_start = c_text.find("{", signature_index)
    if brace_start == -1:
        raise ValueError(f"未找到函数起始大括号: {func_signature}")

    newline = "\r\n" if "\r\n" in c_text else "\n"
    block_lines = [BlockStart, *new_body_lines, BlockEnd, ""]
    new_block = newline + newline.join(block_lines)

    block_start_index = c_text.find(BlockStart, brace_start)
    block_end_index = c_text.find(BlockEnd, brace_start)
    if block_start_index != -1 and block_end_index != -1:
        block_end_index += len(BlockEnd)
        if block_end_index < len(c_text) and c_text[block_end_index:block_end_index + len(newline)] == newline:
            block_end_index += len(newline)
        return c_text[:block_start_index] + new_block + c_text[block_end_index:]

    return c_text[: brace_start + 1] + new_block + c_text[brace_start + 1 :]


c_file_path = script_dir.parent / "src" / "SF020_SysFaultMngt.c"
h_file_path = script_dir.parent / "include" / "SF020_SysFaultMngt.h"
c_text = c_file_path.read_text(encoding="utf-8")
updated_c_text = insert_code_at_function_start(
    c_text,
    "void SF020_SysFaultMngt_Cycle_5ms(void)",
    Merged_Code,
)
updated_c_text = updated_c_text.replace(
    "Rte_Call_DiagnosticMonitor_SetEventStatus",
    "Dem_SetEventStatus",
)
updated_c_text = updated_c_text.replace(
    "DEM_EVENT_STATUS_FAILED",
    "Monr_SF020_u8_DTCIndex_UL, DEM_EVENT_STATUS_FAILED",
)
updated_c_text = updated_c_text.replace(
    "DEM_EVENT_STATUS_PASSED",
    "Monr_SF020_u8_DTCIndex_UL, DEM_EVENT_STATUS_PASSED",
)
updated_c_text = updated_c_text.replace(
    "static void SF020_SysFaultMngt_NoXTCRecover(boolean *rty_Ret);",
    "static void SF020_SysFaultMngt_NoXTCRecover(boolean *rty_Ret);\nvoid RecoverFault(uint8 XtcId, boolean *Ret);\nvoid TriggerFault(uint8 XtcId, uint8 Para, boolean *Ret);",
)
c_file_path.write_text(updated_c_text, encoding="utf-8")
print(f"\n已更新: {c_file_path}")

h_text = h_file_path.read_text(encoding="utf-8")
updated_h_text = h_text.replace(
    '#include "XPSbW_PublicCal.h"',
    '#include "XPSbW_PublicCal.h"\n#include "Dem.h"',
)
h_file_path.write_text(updated_h_text, encoding="utf-8")
print(f"\n已更新: {h_file_path}")


print("\n拼接后的完整代码：")
for code in Merged_Code:
    print(code)
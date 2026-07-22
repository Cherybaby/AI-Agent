from pathlib import Path

from openpyxl import load_workbook


def to_int(value, column_name, row_idx):
	"""Convert cell value to int, supporting decimal and 0x-prefixed strings."""
	if value is None or value == "":
		return 0

	if isinstance(value, bool):
		return int(value)

	if isinstance(value, (int, float)):
		return int(value)

	if isinstance(value, str):
		text = value.strip()
		if not text:
			return 0
		return int(text, 0)

	raise ValueError(
		f"Row {row_idx}: column {column_name} has unsupported type {type(value).__name__}"
	)


def validate_range(value, bits, column_name, row_idx):
	max_value = (1 << bits) - 1
	if value < 0 or value > max_value:
		raise ValueError(
			f"Row {row_idx}: column {column_name} value {value} out of range 0~{max_value}"
		)


def main():
	script_dir = Path(__file__).resolve().parent
	xlsx_path = script_dir.parent / "model" / "SF020_SysFaultMngt.xlsx"

	workbook = load_workbook(filename=xlsx_path, data_only=True)
	if "CalGeneration" not in workbook.sheetnames:
		raise ValueError("Sheet 'CalGeneration' not found in workbook")

	sheet = workbook["CalGeneration"]

	packed_values = []
	seen_a_values = set()

	# Start from row 3 as requested.
	for row_idx in range(3, sheet.max_row + 1):
		a_raw = sheet.cell(row=row_idx, column=1).value
		b_raw = sheet.cell(row=row_idx, column=2).value
		c_raw = sheet.cell(row=row_idx, column=3).value
		d_raw = sheet.cell(row=row_idx, column=4).value
		e_raw = sheet.cell(row=row_idx, column=5).value
		f_raw = sheet.cell(row=row_idx, column=6).value

		# Skip fully empty tail rows.
		if all(v in (None, "") for v in (a_raw, b_raw, c_raw, d_raw, e_raw, f_raw)):
			continue

		a = to_int(a_raw, "A", row_idx)

		# Step 1: if A has duplicate values, keep only the first row.
		if a in seen_a_values:
			continue
		seen_a_values.add(a)

		b = to_int(b_raw, "B", row_idx)
		c = to_int(c_raw, "C", row_idx)
		d = to_int(d_raw, "D", row_idx)
		e = to_int(e_raw, "E", row_idx)
		f = to_int(f_raw, "F", row_idx)

		validate_range(b, 8, "B", row_idx)  # bit0-7
		validate_range(c, 8, "C", row_idx)  # bit8-15
		validate_range(d, 2, "D", row_idx)  # bit16-17
		validate_range(e, 1, "E", row_idx)  # bit18
		validate_range(f, 1, "F", row_idx)  # bit19

		packed = (b & 0xFF) | ((c & 0xFF) << 8) | ((d & 0x3) << 16) | ((e & 0x1) << 18) | ((f & 0x1) << 19)
		packed_values.append(str(packed))

	print(",".join(packed_values))


if __name__ == "__main__":
	main()
